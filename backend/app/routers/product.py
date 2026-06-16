"""
Product and Category router — Phase 3A endpoints.
"""

from typing import List, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status,Form,UploadFile
from sqlalchemy import select, func, or_
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.deps import get_current_user, require_admin
from app.models.product import Product, ProductImage, ProductStatus
from app.models.pricing import VendorPricing, RetailerPricing, DealerPricing
from app.models.user import User, UserRole
from app.schemas.product import (
    BulkPriceUpdate,
    ProductCreate,
    ProductResponse,
    ProductUpdate,
    StockAdjustment,
)
from app.services.audit_service import AuditService
from app.services.category_service import validate_leaf_category
from app.utils.slug import generate_slug, generate_unique_slug

router = APIRouter(tags=["Products & Categories"])


# ── Image Upload ─────────────────────────────────────────────

@router.post("/upload", dependencies=[Depends(get_current_user)])
async def upload_file(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Admin uploads an image and gets static serving URL."""
    import uuid
    import os
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in (".png", ".jpg", ".jpeg", ".webp", ".gif"):
        raise HTTPException(status_code=400, detail="Only images are allowed")
    
    filename = f"{uuid.uuid4()}{ext}"
    os.makedirs("uploads", exist_ok=True)
    filepath = os.path.join("uploads", filename)
    
    with open(filepath, "wb") as f:
        content = await file.read()
        f.write(content)
        
    backend_url = os.getenv("BACKEND_URL", "http://localhost:8000")
    backend_url = backend_url.rstrip("/")
    return {"image_url": f"{backend_url}/uploads/{filename}"}


# ── P3-02: Product CRUD ─────────────────────────────────────

@router.post("/products", response_model=ProductResponse)
async def create_product(
    req: ProductCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Admin creates a product."""
    # Validate GST rate
    if req.gst_rate not in (0, 5, 12, 18, 28):
        raise HTTPException(status_code=400, detail="GST rate must be 0, 5, 12, 18, or 28")

    try:
        await validate_leaf_category(req.category_id, db)
        if req.sub_category_id:
            await validate_leaf_category(req.sub_category_id, db)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    product = Product(
        name=req.name,
        slug=generate_unique_slug(req.name),
        sku=req.sku,
        description=req.description,
        return_policy=req.return_policy,
        return_window_days=req.return_window_days,
        unit=req.unit,
        hsn_code=req.hsn_code,
        base_price=req.base_price,
        gst_rate=req.gst_rate,
        stock_qty=req.stock_qty,
        low_stock_threshold=req.low_stock_threshold,
        category_id=req.category_id,
        sub_category_id=req.sub_category_id,
        status=ProductStatus(req.status),
    )
    db.add(product)
    await db.flush()

    # Add images
    if req.image_urls:
        for i, url in enumerate(req.image_urls):
            img = ProductImage(product_id=product.id, image_url=url, sort_order=i)
            db.add(img)
        await db.flush()

    # Add pricing overrides
    if req.vendor_price is not None:
        db.add(VendorPricing(product_id=product.id, price=req.vendor_price))
    if req.retailer_price is not None:
        db.add(RetailerPricing(product_id=product.id, price=req.retailer_price))
    await db.flush()

    audit = AuditService(db)
    await audit.log_action(current_user, "create_product", "product", product.id)
    # Refresh to load images and pricing relationships
    await db.refresh(product)
    return product


@router.get("/products", response_model=List[ProductResponse])
async def list_products(
    keyword: Optional[str] = None,
    category_id: Optional[UUID] = None,
    sub_category_id: Optional[UUID] = None,
    category: Optional[str] = None,
    sub_category: Optional[str] = None,
    price_min: Optional[int] = None,
    price_max: Optional[int] = None,
    in_stock: Optional[bool] = None,
    sort: Optional[str] = Query(None, pattern="^(price_asc|price_desc|newest|popular)$"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List products with search, filter, sort, and pagination (P3-05)."""
    query = select(Product).where(Product.is_deleted == False, Product.status == ProductStatus.ACTIVE)  # noqa: E712

    if keyword:
        query = query.where(
            or_(
                Product.name.ilike(f"%{keyword}%"),
                Product.description.ilike(f"%{keyword}%"),
                Product.sku.ilike(f"%{keyword}%"),
            )
        )
    if category_id:
        query = query.where(or_(Product.category_id == category_id, Product.sub_category_id == category_id))
    elif category:
        cat_select = select(Category.id).where(or_(Category.slug == category, Category.name.ilike(category)), Category.is_deleted == False)
        cat_id = (await db.execute(cat_select)).scalar_one_or_none()
        if cat_id:
            query = query.where(Product.category_id == cat_id)

    if sub_category_id:
        query = query.where(Product.sub_category_id == sub_category_id)
    elif sub_category:
        subcat_select = select(Category.id).where(or_(Category.slug == sub_category, Category.name.ilike(sub_category)), Category.is_deleted == False)
        subcat_id = (await db.execute(subcat_select)).scalar_one_or_none()
        if subcat_id:
            query = query.where(Product.sub_category_id == subcat_id)
    if price_min is not None:
        query = query.where(Product.base_price >= price_min)
    if price_max is not None:
        query = query.where(Product.base_price <= price_max)
    if in_stock:
        query = query.where(Product.stock_qty > 0)

    # Sort
    if sort == "price_asc":
        query = query.order_by(Product.base_price.asc())
    elif sort == "price_desc":
        query = query.order_by(Product.base_price.desc())
    elif sort == "newest":
        query = query.order_by(Product.created_at.desc())
    else:
        query = query.order_by(Product.created_at.desc())

    # Pagination
    offset = (page - 1) * page_size
    query = query.offset(offset).limit(page_size)

    result = await db.execute(query)
    products = result.scalars().all()

    response_products = []
    for p in products:
        resolved_price = p.base_price
        if current_user.role == UserRole.VENDOR and p.vendor_pricing:
            resolved_price = p.vendor_pricing.price
        elif current_user.role == UserRole.RETAILER and p.retailer_pricing:
            resolved_price = p.retailer_pricing.price

        prod_dict = {
            "id": p.id,
            "name": p.name,
            "slug": p.slug,
            "sku": p.sku,
            "description": p.description,
            "return_policy": p.return_policy,
            "return_window_days": p.return_window_days,
            "unit": p.unit,
            "hsn_code": p.hsn_code,
            "base_price": resolved_price,
            "gst_rate": p.gst_rate,
            "stock_qty": p.stock_qty,
            "low_stock_threshold": p.low_stock_threshold,
            "status": p.status.value,
            "category_id": p.category_id,
            "sub_category_id": p.sub_category_id,
            "images": [{"id": img.id, "image_url": img.image_url, "sort_order": img.sort_order} for img in p.images],
            "vendor_price": p.vendor_price,
            "retailer_price": p.retailer_price
        }
        response_products.append(prod_dict)
    return response_products


@router.get("/products/{product_id}", response_model=ProductResponse)
async def get_product(
    product_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get single product detail."""
    result = await db.execute(
        select(Product).where(Product.id == product_id, Product.is_deleted == False)  # noqa: E712
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    resolved_price = product.base_price
    if current_user.role == UserRole.VENDOR and product.vendor_pricing:
        resolved_price = product.vendor_pricing.price
    elif current_user.role == UserRole.RETAILER and product.retailer_pricing:
        resolved_price = product.retailer_pricing.price

    return {
        "id": product.id,
        "name": product.name,
        "slug": product.slug,
        "sku": product.sku,
        "description": product.description,
        "return_policy": product.return_policy,
        "return_window_days": product.return_window_days,
        "unit": product.unit,
        "hsn_code": product.hsn_code,
        "base_price": resolved_price,
        "gst_rate": product.gst_rate,
        "stock_qty": product.stock_qty,
        "low_stock_threshold": product.low_stock_threshold,
        "status": product.status.value,
        "category_id": product.category_id,
        "sub_category_id": product.sub_category_id,
        "images": [{"id": img.id, "image_url": img.image_url, "sort_order": img.sort_order} for img in product.images],
        "vendor_price": product.vendor_price,
        "retailer_price": product.retailer_price
    }


@router.patch("/products/{product_id}", response_model=ProductResponse)
async def update_product(
    product_id: UUID,
    req: ProductUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Admin updates a product."""
    result = await db.execute(
        select(Product).where(Product.id == product_id, Product.is_deleted == False)  # noqa: E712
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    update_data = req.model_dump(exclude_unset=True)
    if "gst_rate" in update_data and update_data["gst_rate"] not in (0, 5, 12, 18, 28):
        raise HTTPException(status_code=400, detail="GST rate must be 0, 5, 12, 18, or 28")
    if "name" in update_data:
        update_data["slug"] = generate_unique_slug(update_data["name"])
    if "category_id" in update_data:
        try:
            await validate_leaf_category(update_data["category_id"], db)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc))
    if "sub_category_id" in update_data and update_data["sub_category_id"] is not None:
        try:
            await validate_leaf_category(update_data["sub_category_id"], db)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc))

    # Handle image urls update
    if "image_urls" in update_data:
        image_urls = update_data.pop("image_urls")
        from sqlalchemy import delete
        await db.execute(delete(ProductImage).where(ProductImage.product_id == product_id))
        if image_urls:
            for i, url in enumerate(image_urls):
                img = ProductImage(product_id=product_id, image_url=url, sort_order=i)
                db.add(img)
        await db.flush()

    # Handle vendor pricing override
    if "vendor_price" in update_data:
        vp_val = update_data.pop("vendor_price")
        vp_res = await db.execute(select(VendorPricing).where(VendorPricing.product_id == product_id, VendorPricing.is_deleted == False))
        vp = vp_res.scalar_one_or_none()
        if vp_val is not None:
            if vp:
                vp.price = vp_val
            else:
                db.add(VendorPricing(product_id=product_id, price=vp_val))
        else:
            if vp:
                await db.delete(vp)
        await db.flush()

    # Handle retailer pricing override
    if "retailer_price" in update_data:
        rp_val = update_data.pop("retailer_price")
        rp_res = await db.execute(select(RetailerPricing).where(RetailerPricing.product_id == product_id, RetailerPricing.is_deleted == False))
        rp = rp_res.scalar_one_or_none()
        if rp_val is not None:
            if rp:
                rp.price = rp_val
            else:
                db.add(RetailerPricing(product_id=product_id, price=rp_val))
        else:
            if rp:
                await db.delete(rp)
        await db.flush()

    for k, v in update_data.items():
        setattr(product, k, v)
    await db.flush()

    audit = AuditService(db)
    await audit.log_action(current_user, "update_product", "product", product_id, update_data)
    await db.refresh(product)
    return product


# ── P3-03: Role-Based Pricing ────────────────────────────────

@router.patch("/products/bulk-price", dependencies=[Depends(require_admin)])
async def bulk_price_update(
    updates: List[BulkPriceUpdate],
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Admin bulk-updates vendor and retailer pricing."""
    audit = AuditService(db)
    for upd in updates:
        if upd.vendor_price is not None:
            result = await db.execute(
                select(VendorPricing).where(
                    VendorPricing.product_id == upd.product_id, VendorPricing.is_deleted == False  # noqa: E712
                )
            )
            vp = result.scalar_one_or_none()
            if vp:
                vp.price = upd.vendor_price
            else:
                db.add(VendorPricing(product_id=upd.product_id, price=upd.vendor_price))

        if upd.retailer_price is not None:
            result = await db.execute(
                select(RetailerPricing).where(
                    RetailerPricing.product_id == upd.product_id, RetailerPricing.is_deleted == False  # noqa: E712
                )
            )
            rp = result.scalar_one_or_none()
            if rp:
                rp.price = upd.retailer_price
            else:
                db.add(RetailerPricing(product_id=upd.product_id, price=upd.retailer_price))

    await db.flush()
    await audit.log_action(current_user, "bulk_price_update", "pricing", diff_json={"count": len(updates)})
    return {"message": f"Updated pricing for {len(updates)} products"}


# ── P3-06: Stock Adjustment ──────────────────────────────────

@router.patch("/products/{product_id}/stock")
async def adjust_stock(
    product_id: UUID,
    req: StockAdjustment,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Admin manually adjusts product stock."""
    result = await db.execute(
        select(Product).where(Product.id == product_id, Product.is_deleted == False)  # noqa: E712
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    new_qty = product.stock_qty + req.adjustment
    if new_qty < 0:
        raise HTTPException(status_code=400, detail="Stock cannot go below 0")

    old_qty = product.stock_qty
    product.stock_qty = new_qty
    await db.flush()

    audit = AuditService(db)
    await audit.log_action(
        current_user, "stock_adjustment", "product", product_id,
        {"old_qty": old_qty, "new_qty": new_qty, "adjustment": req.adjustment, "reason": req.reason},
    )
    return {"product_id": str(product_id), "old_stock": old_qty, "new_stock": new_qty}


# ── P3-07: Excel/CSV Bulk Upload ─────────────────────────────

@router.post("/products/bulk-upload", dependencies=[Depends(require_admin)])
async def bulk_upload_products(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Admin bulk-uploads products from CSV or XLSX."""
    import io
    import csv
    from openpyxl import load_workbook

    contents = await file.read()
    rows = []

    if file.filename.endswith(".csv"):
        try:
            text = contents.decode("utf-8")
            reader = csv.reader(io.StringIO(text))
            rows_raw = list(reader)
            if not rows_raw:
                raise HTTPException(status_code=400, detail="Empty file")
            headers = rows_raw[0]
            for i in range(1, len(rows_raw)):
                if not any(rows_raw[i]):  # skip empty lines
                    continue
                row_dict = {}
                for idx, h in enumerate(headers):
                    if idx < len(rows_raw[i]):
                        row_dict[h] = rows_raw[i][idx]
                rows.append((i + 1, row_dict))
        except Exception as e:
            if isinstance(e, HTTPException):
                raise e
            raise HTTPException(status_code=400, detail=f"Failed to parse CSV: {str(e)}")

    elif file.filename.endswith((".xlsx", ".xls")):
        try:
            wb = load_workbook(filename=io.BytesIO(contents), read_only=True)
            sheet = wb.active
            rows_raw = list(sheet.iter_rows(values_only=True))
            if not rows_raw:
                raise HTTPException(status_code=400, detail="Empty file")
            headers = [str(h) if h is not None else "" for h in rows_raw[0]]
            for i in range(1, len(rows_raw)):
                row_vals = rows_raw[i]
                if not any(v is not None for v in row_vals):  # skip empty lines
                    continue
                row_dict = {}
                for idx, h in enumerate(headers):
                    if h and idx < len(row_vals):
                        row_dict[h] = row_vals[idx]
                rows.append((i + 1, row_dict))
        except Exception as e:
            if isinstance(e, HTTPException):
                raise e
            raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file format. Please upload .csv or .xlsx")

    # Map headers to standard field names
    HEADER_MAP = {
        "name": "name",
        "sku": "sku",
        "description": "description",
        "unit": "unit",
        "hsncode": "hsn_code",
        "baseprice": "base_price",
        "price": "base_price",
        "gstrate": "gst_rate",
        "gst": "gst_rate",
        "stockqty": "stock_qty",
        "stock": "stock_qty",
        "inventory": "stock_qty",
        "lowstockthreshold": "low_stock_threshold",
        "threshold": "low_stock_threshold",
        "category": "category_name",
        "categoryname": "category_name",
    }

    def normalize_header(h: str) -> str:
        return h.strip().lower().replace("_", "").replace(" ", "")

    errors = []
    success_count = 0
    file_skus = set()

    for row_num, raw_row in rows:
        row_errors = []
        
        # Standardize the dict
        data = {}
        for k, v in raw_row.items():
            norm_k = normalize_header(k)
            if norm_k in HEADER_MAP:
                data[HEADER_MAP[norm_k]] = v

        # Validate required fields
        name = str(data.get("name") or "").strip()
        sku = str(data.get("sku") or "").strip()
        category_name = str(data.get("category_name") or "").strip()

        if not name:
            row_errors.append("Product name is required")
        if not sku:
            row_errors.append("SKU is required")
        if not category_name:
            row_errors.append("Category is required")

        # Validate/convert numbers
        base_price_val = data.get("base_price")
        base_price = 0
        if base_price_val is None:
            row_errors.append("Base price is required")
        else:
            try:
                # Convert from decimal (Rupees) to paise (integer)
                base_price = int(round(float(base_price_val) * 100))
                if base_price <= 0:
                    row_errors.append("Base price must be greater than 0")
            except (ValueError, TypeError):
                row_errors.append(f"Invalid base price format: {base_price_val}")

        gst_val = data.get("gst_rate")
        gst_rate = 18
        if gst_val is not None:
            try:
                gst_rate = int(gst_val)
                if gst_rate not in (0, 5, 12, 18, 28):
                    row_errors.append(f"Invalid GST rate: {gst_rate}. Must be 0, 5, 12, 18, or 28")
            except (ValueError, TypeError):
                row_errors.append(f"Invalid GST rate format: {gst_val}")

        stock_val = data.get("stock_qty")
        stock_qty = 0
        if stock_val is not None:
            try:
                stock_qty = int(stock_val)
                if stock_qty < 0:
                    row_errors.append("Stock quantity cannot be negative")
            except (ValueError, TypeError):
                row_errors.append(f"Invalid stock qty format: {stock_val}")

        threshold_val = data.get("low_stock_threshold")
        low_stock_threshold = 10
        if threshold_val is not None:
            try:
                low_stock_threshold = int(threshold_val)
                if low_stock_threshold < 0:
                    row_errors.append("Low stock threshold cannot be negative")
            except (ValueError, TypeError):
                row_errors.append(f"Invalid low stock threshold format: {threshold_val}")

        # Check SKU uniqueness
        if sku:
            if sku in file_skus:
                row_errors.append(f"Duplicate SKU '{sku}' within the uploaded file")
            else:
                file_skus.add(sku)
                # Check DB
                db_sku_check = await db.execute(select(Product).where(Product.sku == sku, Product.is_deleted == False))
                if db_sku_check.scalar_one_or_none():
                    row_errors.append(f"SKU '{sku}' already exists in database")

        if row_errors:
            errors.append(f"Row {row_num}: {', '.join(row_errors)}")
            continue

        # Find or create category
        cat_result = await db.execute(
            select(Category).where(func.lower(Category.name) == category_name.lower(), Category.is_deleted == False)
        )
        cat = cat_result.scalar_one_or_none()
        if not cat:
            cat = Category(
                name=category_name,
                slug=generate_unique_slug(category_name),
                visible_to_vendor=True,
                visible_to_retailer=True,
            )
            db.add(cat)
            await db.flush()

        # Create product
        product = Product(
            name=name,
            slug=generate_unique_slug(name),
            sku=sku,
            description=str(data.get("description") or "").strip() or None,
            unit=str(data.get("unit") or "piece").strip(),
            hsn_code=str(data.get("hsn_code") or "").strip() or None,
            base_price=base_price,
            gst_rate=gst_rate,
            stock_qty=stock_qty,
            low_stock_threshold=low_stock_threshold,
            category_id=cat.id,
            status=ProductStatus.ACTIVE,
        )
        db.add(product)
        success_count += 1

    if errors:
        # If there are errors, rollback database session and raise error
        await db.rollback()
        raise HTTPException(
            status_code=400,
            detail={"message": "Validation failed during bulk upload", "errors": errors}
        )

    await db.flush()
    audit = AuditService(db)
    await audit.log_action(current_user, "bulk_upload_products", "product", diff_json={"count": success_count})
    return {"message": f"Successfully imported {success_count} products", "success_count": success_count}
